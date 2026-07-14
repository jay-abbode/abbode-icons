#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Abbode Icon Match Preview
=========================
A faithful mini-run of the Shopify <-> App icon matcher on a curated sample.

Per Shopify (old) icon:
  1) COLOR-INVARIANT RANKER: grayscale the icon, embed with DINOv2, take the
     top-K nearest app icons by cosine similarity. Grayscale means a recolored
     design still retrieves its own shape family (red heart -> the heart family).
  2) VISION CONFIRM: show Claude the OLD icon and the K candidates *in color*,
     and let it pick the true design match. Because the confirm step sees color,
     it (a) treats a recolor as a match, (b) picks the right color when the app
     splits a shape by color, and (c) reads writing to separate shared shells
     (Hot Dog vs New York Hot Dog, the four locket variants, etc.).

Outputs (next to this script):
  match_preview.html  -> self-contained visual report: old -> new pairs, each
                         with a confidence badge, the reason, and the runner-up
                         candidates so you can inspect near-misses.
  match_preview.csv   -> the exact column shape the full run will produce.

This is a PREVIEW. The full build adds: live Shopify pagination (no baked
sample), the reverse bucket (app icons with no Shopify match = new additions),
and the metaobject mapping. Nothing here is throwaway - it's the real engine.
"""

import os
import sys
import re
import io
import json
import base64
import hashlib
import html
import time
import traceback
from pathlib import Path

# --------------------------------------------------------------------------- #
# Config
# --------------------------------------------------------------------------- #
HERE = Path(__file__).resolve().parent
SAMPLE_JSON = HERE / "match_preview_sample.json"  # legacy; unused when an export CSV is present
CACHE = HERE / ".cache-match"
IMG_CACHE = CACHE / "img"
EMB_CACHE = CACHE / "emb"
CONF_CACHE = CACHE / "conf"
IMG_DIR = HERE / "match_imgs"          # thumbnails written here, referenced relatively by the HTML
OUT_HTML = HERE / "icon_review.html"
OUT_CSV = HERE / "icon_match_auto.csv"

# Icon catalog (source of truth) - MASTER tab, header row 2, data row 3+
SHEET_ID = "1zP1wTjPpYxhEQ4GnF8pCLdZj1DiyoGIrwEr3VWrLbqo"
SHEET_RANGE = "MASTER!A2:AB"
COL_CATEGORY, COL_NAME, COL_STATUS, COL_PNG = 0, 2, 7, 13  # A, C, H, N

MODEL = os.environ.get("MATCH_MODEL", "claude-sonnet-5")   # any vision-capable model
TOP_K = int(os.environ.get("MATCH_TOP_K", "12"))           # candidates shown to the confirm step
IMG_SIZE = int(os.environ.get("MATCH_IMG_SIZE", "224"))    # DINOv2 input (multiple of 14)
APP_LIMIT = int(os.environ.get("MATCH_APP_LIMIT", "0"))    # 0 = whole catalog
APP_ONLY_LIMIT = int(os.environ.get("MATCH_APP_ONLY_LIMIT", "0"))  # app-only cards shown (0 = all)
CONFIRM_WORKERS = int(os.environ.get("MATCH_CONFIRM_WORKERS", "8"))  # parallel Claude calls
PROMPT_VERSION = "v2-strict"   # bump to re-judge cached decisions after prompt changes

IMAGENET_MEAN = [0.485, 0.456, 0.406]
IMAGENET_STD = [0.229, 0.224, 0.225]

BRAND = {
    "porcelain": "#FFFCF7", "ink": "#432222", "pink": "#F2B2AE", "grey": "#6E6E6E",
    "good": "#3f7d5a", "warn": "#b8863b", "bad": "#a24b4b", "line": "#EAE2D6",
}

for _d in (IMG_CACHE, EMB_CACHE, CONF_CACHE, IMG_DIR):
    _d.mkdir(parents=True, exist_ok=True)


def log(*a):
    print("[match]", *a, flush=True)


# --------------------------------------------------------------------------- #
# Credentials (mirrors the app's loading order)
# --------------------------------------------------------------------------- #
def google_credentials():
    from google.oauth2 import service_account
    scopes = [
        "https://www.googleapis.com/auth/spreadsheets.readonly",
        "https://www.googleapis.com/auth/drive.readonly",
    ]
    for p in (Path.cwd() / "google-credentials.json", HERE / "google-credentials.json"):
        if p.exists():
            log(f"Using Google credentials: {p}")
            return service_account.Credentials.from_service_account_file(str(p), scopes=scopes)
    envp = os.environ.get("GOOGLE_APPLICATION_CREDENTIALS")
    if envp and Path(envp).exists():
        log(f"Using Google credentials: {envp}")
        return service_account.Credentials.from_service_account_file(envp, scopes=scopes)
    email = os.environ.get("GOOGLE_SERVICE_ACCOUNT_EMAIL")
    key = os.environ.get("GOOGLE_SERVICE_ACCOUNT_PRIVATE_KEY")
    if email and key:
        log("Using Google credentials from GOOGLE_SERVICE_ACCOUNT_* env vars")
        info = {
            "type": "service_account",
            "client_email": email,
            "private_key": key.replace("\\n", "\n"),
            "token_uri": "https://oauth2.googleapis.com/token",
        }
        return service_account.Credentials.from_service_account_info(info, scopes=scopes)
    sys.exit(
        "ERROR: No Google credentials found.\n"
        "  Put google-credentials.json in this folder (or the abbode-icons repo root),\n"
        "  or set GOOGLE_APPLICATION_CREDENTIALS to its path,\n"
        "  or set GOOGLE_SERVICE_ACCOUNT_EMAIL and GOOGLE_SERVICE_ACCOUNT_PRIVATE_KEY."
    )


# --------------------------------------------------------------------------- #
# App catalog from the MASTER sheet
# --------------------------------------------------------------------------- #
FILEID_RE = re.compile(r"(?:[?&]id=|/d/|/file/d/)([A-Za-z0-9_-]{20,})")


def load_shopify_catalog():
    """Read the live Shopify icon catalog from a product export CSV.

    In Shopify admin: Products -> Export -> All products -> CSV. Drop the file in this
    folder (any name containing 'products_export' works) or set SHOPIFY_EXPORT_CSV.
    We group rows by Handle, keep products whose SKU starts with 'ICON-', and take the
    Image Position 1 image as the icon.
    """
    import csv as _csv
    import glob as _glob
    path = os.environ.get("SHOPIFY_EXPORT_CSV", "").strip()
    if not path:
        found = sorted(_glob.glob(str(HERE / "*products_export*.csv"))
                       + _glob.glob(str(HERE / "shopify_products*.csv"))
                       + _glob.glob(str(Path.cwd() / "*products_export*.csv")))
        path = found[0] if found else ""
    if not path or not Path(path).exists():
        sys.exit(
            "ERROR: No Shopify product export found.\n"
            "  In Shopify admin: Products -> Export -> All products (or Current search) -> CSV,\n"
            "  then drop that .csv into this folder (any name with 'products_export' works),\n"
            "  or set SHOPIFY_EXPORT_CSV to its full path."
        )
    log(f"Reading Shopify catalog from export: {Path(path).name}")

    def col(row, *names):
        for n in names:
            for k in row:
                if k and k.strip().lower() == n.lower():
                    return (row[k] or "").strip()
        return ""

    prods = {}
    with open(path, newline="", encoding="utf-8-sig") as f:
        for row in _csv.DictReader(f):
            handle = col(row, "Handle")
            if not handle:
                continue
            p = prods.setdefault(handle, {"handle": handle, "title": "", "sku": "",
                                          "image_url": "", "img_pos": None})
            title = col(row, "Title")
            if title and not p["title"]:
                p["title"] = title
            sku = col(row, "Variant SKU")
            if sku and not p["sku"]:
                p["sku"] = sku
            img = col(row, "Image Src")
            if img:
                pos = col(row, "Image Position")
                try:
                    posn = int(pos) if pos else 999
                except ValueError:
                    posn = 999
                if p["img_pos"] is None or posn < p["img_pos"]:
                    p["img_pos"] = posn
                    p["image_url"] = img

    out = []
    for p in prods.values():
        if not p["sku"].upper().startswith("ICON-") or not p["image_url"]:
            continue
        out.append({"gid": "", "handle": p["handle"], "old_name": p["title"] or p["handle"],
                    "sku": p["sku"], "image_url": p["image_url"]})
    out.sort(key=lambda r: r["old_name"].lower())
    log(f"Shopify icons in export (SKU ICON-*): {len(out)}")
    if not out:
        sys.exit("ERROR: No icon products (SKU starting 'ICON-') found in the export CSV.")
    return out


def read_app_catalog(creds):
    """Read the MASTER tab via an xlsx export.

    Column N ("PNG") stores links two ways: most icons are RICH-TEXT hyperlinks
    (a filename with a link attached) and a few are =HYPERLINK() formulas. The
    Sheets values API only exposes the formula ones - that's why an earlier
    version saw ~37 of ~785 icons. Exporting to xlsx and reading each cell's
    hyperlink target (with the raw value as fallback) recovers all of them.
    """
    import io as _io
    import warnings
    import openpyxl
    from googleapiclient.discovery import build

    drive = build("drive", "v3", credentials=creds, cache_discovery=False)
    data = drive.files().export(
        fileId=SHEET_ID,
        mimeType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    ).execute()

    with warnings.catch_warnings():
        warnings.simplefilter("ignore")  # ignore openpyxl pivot-cache warnings
        wb = openpyxl.load_workbook(_io.BytesIO(data), data_only=False)
    ws = wb["MASTER"] if "MASTER" in wb.sheetnames else wb[wb.sheetnames[0]]

    name_col = COL_NAME + 1       # openpyxl is 1-based
    png_col = COL_PNG + 1
    status_col = COL_STATUS + 1
    cat_col = COL_CATEGORY + 1

    out, seen, skipped = [], set(), 0
    for r in range(3, ws.max_row + 1):  # data starts at row 3 (header is row 2)
        nv = ws.cell(r, name_col).value
        name = str(nv).strip() if nv is not None else ""

        pcell = ws.cell(r, png_col)
        if pcell.hyperlink and pcell.hyperlink.target:
            link = pcell.hyperlink.target          # rich-text hyperlink (most rows)
        elif pcell.value is not None:
            link = str(pcell.value)                # =HYPERLINK(...) or plain URL
        else:
            link = ""

        m = FILEID_RE.search(link)
        if not name or not m:
            skipped += 1
            continue
        fid = m.group(1)
        if fid in seen:
            continue
        seen.add(fid)
        sv = ws.cell(r, status_col).value
        cv = ws.cell(r, cat_col).value
        out.append({
            "name": name,
            "status": str(sv).strip() if sv is not None else "",
            "category": str(cv).strip() if cv is not None else "",
            "fileid": fid,
        })
    wb.close()

    out.sort(key=lambda a: a["name"].lower())
    log(f"App catalog: {len(out)} icons with PNG links ({skipped} rows skipped - no name/link)")
    if APP_LIMIT > 0:
        out = out[:APP_LIMIT]
        log(f"MATCH_APP_LIMIT set -> using first {len(out)} app icons")
    return out


# --------------------------------------------------------------------------- #
# Image fetch (cached to disk so re-runs are instant and Claude is never re-billed)
# --------------------------------------------------------------------------- #
def _cache_path(base, key, ext):
    return base / (hashlib.sha256(key.encode()).hexdigest()[:24] + ext)


def fetch_shopify_image(url):
    import requests
    p = _cache_path(IMG_CACHE, "shopify:" + url, ".png")
    if p.exists():
        return p.read_bytes()
    r = requests.get(url, timeout=60)
    r.raise_for_status()
    p.write_bytes(r.content)
    return r.content


def make_drive_session(creds):
    from google.auth.transport.requests import AuthorizedSession
    return AuthorizedSession(creds)


def fetch_drive_image(session, fileid):
    # supportsAllDrives=true is REQUIRED for the shared-drive PNG folder,
    # otherwise Drive returns a silent failure for shared-drive files.
    p = _cache_path(IMG_CACHE, "drive:" + fileid, ".png")
    if p.exists():
        return p.read_bytes()
    url = (
        f"https://www.googleapis.com/drive/v3/files/{fileid}"
        f"?alt=media&supportsAllDrives=true"
    )
    r = session.get(url, timeout=60)
    r.raise_for_status()
    p.write_bytes(r.content)
    return r.content


# --------------------------------------------------------------------------- #
# Image ops (PIL)
# --------------------------------------------------------------------------- #
def open_rgba(data):
    from PIL import Image
    return Image.open(io.BytesIO(data)).convert("RGBA")


def on_white(img):
    """Composite transparent PNG onto white -> RGB (color preserved)."""
    from PIL import Image
    bg = Image.new("RGBA", img.size, (255, 255, 255, 255))
    bg.alpha_composite(img)
    return bg.convert("RGB")


def letterbox(img, size):
    """Pad to square (white) then resize - preserves aspect ratio."""
    from PIL import Image
    w, h = img.size
    s = max(w, h, 1)
    canvas = Image.new("RGB", (s, s), (255, 255, 255))
    canvas.paste(img, ((s - w) // 2, (s - h) // 2))
    return canvas.resize((size, size), Image.LANCZOS)


def gray3(img):
    """Grayscale replicated to 3 channels -> hue removed, structure kept."""
    from PIL import Image
    g = img.convert("L")
    return Image.merge("RGB", (g, g, g))


def data_uri(img, maxdim):
    im = img.copy()
    w, h = im.size
    if max(w, h) > maxdim:
        sc = maxdim / max(w, h)
        im = im.resize((max(1, int(w * sc)), max(1, int(h * sc))))
    buf = io.BytesIO()
    im.save(buf, format="PNG")
    return "data:image/png;base64," + base64.b64encode(buf.getvalue()).decode()


def b64_png(img, maxdim=320):
    im = img.copy()
    w, h = im.size
    if max(w, h) > maxdim:
        sc = maxdim / max(w, h)
        im = im.resize((max(1, int(w * sc)), max(1, int(h * sc))))
    buf = io.BytesIO()
    im.save(buf, format="PNG")
    return base64.b64encode(buf.getvalue()).decode()


def safe_key(s):
    return re.sub(r"[^A-Za-z0-9_-]", "_", str(s))[:80]


def save_img(key, pil, size=150):
    """Write a square thumbnail to match_imgs/ (once) and return its relative path.
    Keeps the HTML tiny so hundreds of icons load smoothly (lazy from disk)."""
    if pil is None:
        return ""
    p = IMG_DIR / (key + ".png")
    if not p.exists():
        try:
            letterbox(pil, size).save(p, format="PNG")
        except Exception:
            return ""
    return IMG_DIR.name + "/" + key + ".png"


# --------------------------------------------------------------------------- #
# DINOv2 embedder (grayscale, aspect-preserving), cached per image content
# --------------------------------------------------------------------------- #
class Embedder:
    def __init__(self):
        import torch
        import timm
        import numpy as np
        self.torch = torch
        self.np = np
        name = "vit_small_patch14_dinov2.lvd142m"
        try:
            self.model = timm.create_model(
                name, pretrained=True, num_classes=0, dynamic_img_size=True
            )
            self.size = IMG_SIZE
        except Exception:
            log("timm dynamic_img_size unavailable -> native 518 input (slower).")
            self.model = timm.create_model(name, pretrained=True, num_classes=0)
            self.size = 518
        self.model.eval()
        self.mean = torch.tensor(IMAGENET_MEAN).view(1, 3, 1, 1)
        self.std = torch.tensor(IMAGENET_STD).view(1, 3, 1, 1)
        log(f"DINOv2 loaded ({name}, input {self.size}px).")

    def _tensor(self, pil_rgb_sized):
        arr = self.np.asarray(pil_rgb_sized).astype("float32") / 255.0
        t = self.torch.from_numpy(arr).permute(2, 0, 1).unsqueeze(0)
        return (t - self.mean) / self.std

    def embed(self, data):
        h = hashlib.sha256(data).hexdigest()[:24]
        cp = EMB_CACHE / (h + ".npy")
        if cp.exists():
            return self.np.load(cp)
        sized = letterbox(gray3(on_white(open_rgba(data))), self.size)
        with self.torch.no_grad():
            v = self.model(self._tensor(sized)).squeeze(0).cpu().numpy().astype("float32")
        v = v / (self.np.linalg.norm(v) + 1e-8)
        self.np.save(cp, v)
        return v


# --------------------------------------------------------------------------- #
# Claude vision confirm
# --------------------------------------------------------------------------- #
CONFIRM_SYS = (
    "You are doing STRICT quality-control matching of embroidery icons. An OLD icon "
    "(from a website) must be matched to the SAME icon in a NEW library. Your job is to "
    "find the SAME design - not a similar-looking one.\n\n"
    "THE ONE RULE THAT MATTERS MOST - match the CONTENTS, not the frame.\n"
    "Many icons share an outer shape or frame - lockets, stamps, badges, banners, "
    "circles, squares, shields - but contain DIFFERENT things inside (an animal, a "
    "figure, initials, a word, a symbol). Icons with the same frame but different "
    "contents are DIFFERENT icons and MUST NOT be matched.\n"
    "  - A locket with a BULLDOG inside is NOT the same as a plain/empty locket, a "
    "locket with a bow, or a locket with a different animal. Same frame, different "
    "contents = NOT a match.\n"
    "  - A stamp reading 'BOSTON' is NOT the same as a stamp reading 'CHICAGO'.\n"
    "  - A pennant with one word is NOT the same as a pennant with another word.\n\n"
    "The ONLY thing allowed to differ between the old icon and its true match is COLOR. "
    "Everything else - the subject, the interior details, the layout, any text - must be "
    "the SAME.\n\n"
    "It is common that the correct match is simply NOT among the candidates (the icon may "
    "have been removed). Do NOT force a match. When in doubt, return -1: a false match is "
    "worse than a missed one.\n\n"
    "Process: FIRST look carefully at what the OLD icon actually depicts (what is inside / "
    "what it shows). THEN compare each candidate on that. Pick the candidate that is the "
    "SAME design (same subject and details, color aside), or -1 if none clearly is.\n\n"
    "Respond with ONLY a JSON object, no other text, no code fences:\n"
    '{"old_depicts": "<short: what the OLD icon shows / what is inside it>", '
    '"match_index": <int or -1>, "same_subject": <true|false>, '
    '"confidence": "high"|"medium"|"low", "reason": "<one short sentence>"}'
)

_MODEL_HINT_SHOWN = False


def confirm(client, old_img_color, old_name, cands):
    """cands: list of dicts {name, img_color}. Returns decision dict."""
    global _MODEL_HINT_SHOWN
    key_src = (PROMPT_VERSION + "||" + old_name + "||"
               + "||".join(c["name"] for c in cands) + "||" + MODEL)
    cp = _cache_path(CONF_CACHE, key_src, ".json")
    if cp.exists():
        try:
            return json.loads(cp.read_text())
        except Exception:
            pass

    content = [
        {"type": "text",
         "text": (f"OLD icon (from the website). Old name: '{old_name}'.\n"
                  f"Below are {len(cands)} candidate icons from the NEW library, "
                  f"numbered 0-{len(cands) - 1}, each with its new name.")},
        {"type": "image",
         "source": {"type": "base64", "media_type": "image/png",
                    "data": b64_png(old_img_color)}},
    ]
    for i, c in enumerate(cands):
        content.append({"type": "text", "text": f"Candidate {i} - new name: '{c['name']}'"})
        content.append({"type": "image",
                        "source": {"type": "base64", "media_type": "image/png",
                                   "data": b64_png(c["img_color"])}})
    content.append({"type": "text",
                    "text": ("First state what the OLD icon depicts, then decide which "
                             "candidate is the SAME design (same contents; only color may "
                             "differ). If none is clearly the same, use -1. "
                             "Reply with ONLY the JSON object.")})

    try:
        msg = client.messages.create(
            model=MODEL, max_tokens=400, system=CONFIRM_SYS,
            messages=[{"role": "user", "content": content}],
        )
        txt = "".join(b.text for b in msg.content if getattr(b, "type", "") == "text").strip()
        mjson = re.search(r"\{.*\}", txt, re.S)  # tolerate any stray prose
        raw = json.loads(mjson.group(0)) if mjson else {}

        mi = int(raw.get("match_index", -1))
        same = bool(raw.get("same_subject", False))
        conf = str(raw.get("confidence", "low")).lower()
        if conf not in ("high", "medium", "low"):
            conf = "low"
        depicts = str(raw.get("old_depicts", "")).strip()[:140]
        reason = str(raw.get("reason", "")).strip()[:200]

        # Hard gate: if the subject doesn't match, it is NOT a match, no matter what
        # index was returned. This is what stops "bulldog locket" -> "plain locket".
        if not same or mi < 0 or mi >= len(cands):
            mi = -1

        disp = (f"depicts: {depicts}" if depicts else "")
        if reason:
            disp = (disp + " \u2014 " + reason) if disp else reason
        data = {"match_index": mi, "confidence": conf, "same_shape": same, "reason": disp}
        cp.write_text(json.dumps(data))  # cache ONLY successful judgments
        return data
    except Exception as e:
        if not _MODEL_HINT_SHOWN:
            _MODEL_HINT_SHOWN = True
            emsg = str(e).lower()
            log(f"Vision confirm error: {type(e).__name__}: {e}")
            if "model" in emsg or "not_found" in emsg or "404" in emsg:
                log(f"  -> '{MODEL}' may be wrong for your account. "
                    f"Set a valid model:  setx MATCH_MODEL \"claude-...\"")
            if "api_key" in emsg or "authentication" in emsg or "401" in emsg:
                log("  -> Check ANTHROPIC_API_KEY.")
        # On failure, flag for review rather than assert a match; do NOT cache,
        # so a transient error retries on the next run instead of sticking.
        return {"match_index": -1, "confidence": "low", "same_shape": False,
                "reason": f"vision confirm failed ({type(e).__name__}) - review manually"}


# --------------------------------------------------------------------------- #
# HTML report
# --------------------------------------------------------------------------- #
def esc(s):
    return html.escape(str(s))


def clean_name(s):
    return re.sub(r"\s*Embroidered Icon$", "", str(s)).strip()


def pill(conf, bucket):
    color = {"High confidence": BRAND["good"],
             "Needs review": BRAND["warn"],
             "No match - likely cut": BRAND["bad"]}[bucket]
    label = bucket if bucket == "No match - likely cut" else f"{bucket} - {conf}"
    return (f'<span style="display:inline-block;padding:3px 10px;border-radius:999px;'
            f'font-size:12px;font-weight:600;color:#fff;background:{color}">{esc(label)}</span>')


# Designation options shown as one-click buttons on every pair.
TAG_DEFS = [
    ("keep", "Keep as-is", "#6E6E6E"),
    ("color", "Color change", "#C9748F"),
    ("name", "Name change", "#4E7CA1"),
    ("color_name", "Color + Name", "#8A6BB0"),
    ("cut", "Cut", "#A24B4B"),
    ("wrong", "\u2717 Wrong match", "#432222"),
    ("missed", "Missed", "#B8863B"),
    ("other", "Other", "#4E8A78"),
]
TAG_COLORS = {t[0]: t[2] for t in TAG_DEFS}


def default_tag(r):
    """Smart starting designation so the user mostly confirms, not clicks."""
    if not r["matched"]:
        return "cut"
    if clean_name(r["old_name"]).lower() != clean_name(r["matched"]["name"]).lower():
        return "name"
    return "keep"


def render_html(rows, counts, meta, app_only=None):
    app_only = app_only or []
    for i, r in enumerate(rows):
        r["_id"] = i

    def _autoidx(r):
        for k, c in enumerate(r["cands_view"]):
            if c.get("picked"):
                return k
        return -1

    row_js = []
    for r in rows:
        row_js.append({
            "id": r["_id"],
            "gid": r.get("gid", ""),
            "handle": r.get("handle", ""),
            "sku": r.get("sku", ""),
            "old_name": r["old_name"],
            "confidence": r["conf"],
            "bucket": r["bucket"],
            "defaultTag": default_tag(r),
            "autoIdx": _autoidx(r),
            "cands": [{"name": c["name"], "fileid": c.get("fileid", ""),
                       "category": c.get("category", ""), "status": c.get("status", "")}
                      for c in r["cands_view"]],
        })

    css_branded = ("""
    * { box-sizing: border-box; }
    body { margin:0; padding:28px 32px; background:%(porcelain)s; color:%(ink)s;
           font-family: 'Segoe UI', -apple-system, system-ui, sans-serif; }
    h1 { font-family: Georgia, 'Times New Roman', serif; font-weight:600;
         font-size:26px; margin:0 0 4px; letter-spacing:.2px; }
    .sub { color:%(grey)s; font-size:14px; margin-bottom:6px; }
    .summary { display:flex; gap:10px; flex-wrap:wrap; margin:14px 0 4px; }
    .stat { background:#fff; border:1px solid %(line)s; border-radius:12px;
            padding:10px 16px; min-width:120px; }
    .stat .n { font-size:22px; font-weight:700; }
    .stat .l { font-size:11px; color:%(grey)s; text-transform:uppercase; letter-spacing:.5px; }
    h2 { font-family: Georgia, serif; font-size:17px; margin:24px 0 12px;
         padding-bottom:6px; border-bottom:1px solid %(line)s; }
    .card { background:#fff; border:1px solid %(line)s; border-radius:14px;
            padding:16px; margin-bottom:12px; }
    .pair { display:flex; align-items:center; gap:18px; }
    .thumb { width:104px; height:104px; border-radius:10px; border:1px solid %(line)s;
             background:%(porcelain)s; object-fit:contain; padding:6px; flex:0 0 auto; }
    .col { min-width:150px; }
    .lbl { font-size:11px; color:%(grey)s; text-transform:uppercase; letter-spacing:.5px; }
    .nm { font-size:15px; font-weight:600; margin-top:2px; }
    .meta2 { font-size:12px; color:%(grey)s; margin-top:2px; }
    .arrow { font-size:22px; color:%(pink)s; flex:0 0 auto; }
    .right { margin-left:auto; text-align:right; flex:0 0 auto; }
    .reason { font-size:13px; color:%(ink)s; margin-top:10px; font-style:italic; opacity:.85; }
    .cands { margin-top:12px; display:flex; gap:8px; flex-wrap:wrap; align-items:flex-start; }
    .cwrap { text-align:center; width:78px; }
    .cimg { width:64px; height:64px; border-radius:8px; border:1px solid %(line)s;
            background:%(porcelain)s; object-fit:contain; padding:3px; }
    .cimg.pick { border:2px solid %(good)s; }
    .cname { font-size:10px; color:%(grey)s; margin-top:3px; line-height:1.2;
             overflow:hidden; display:-webkit-box; -webkit-line-clamp:2; -webkit-box-orient:vertical; }
    .csim { font-size:10px; color:%(ink)s; font-weight:600; }
    .none { color:%(bad)s; font-weight:600; font-size:15px; }
    .foot { color:%(grey)s; font-size:12px; margin-top:26px; line-height:1.5; }
    """) % BRAND

    css_static = """
    .help { font-size:13px; color:#6E6E6E; background:#fff; border:1px solid #EAE2D6;
            border-radius:10px; padding:10px 14px; margin:12px 0 6px; }
    .toolbar { position:sticky; top:0; z-index:20; background:#FFFCF7;
               border-bottom:1px solid #EAE2D6; padding:12px 0; margin:14px 0 18px;
               display:flex; flex-wrap:wrap; gap:10px 18px; align-items:center; }
    .counts { display:flex; gap:6px; flex-wrap:wrap; }
    .cbadge { color:#fff; font-size:11px; font-weight:600; padding:3px 9px; border-radius:999px; }
    .filters { display:flex; gap:6px; flex-wrap:wrap; }
    .filters button { font-size:12px; padding:5px 10px; border:1px solid #EAE2D6;
                      background:#fff; color:#432222; border-radius:8px; cursor:pointer; }
    .filters button.fon { background:#432222; color:#fff; border-color:#432222; }
    .dl { margin-left:auto; display:flex; align-items:center; gap:12px; }
    #progress { font-size:12px; color:#6E6E6E; white-space:nowrap; }
    .dlbtn { font-size:13px; font-weight:700; padding:8px 16px; border:none; border-radius:8px;
             background:#432222; color:#fff; cursor:pointer; }
    .dlbtn:hover { opacity:.92; }
    .controls { margin-top:14px; padding-top:12px; border-top:1px solid #EAE2D6;
                display:flex; align-items:center; gap:8px; flex-wrap:wrap; }
    .tagbtn { font-size:12px; font-weight:600; padding:6px 12px; border:1.5px solid;
              border-radius:999px; background:transparent; cursor:pointer; transition:all .08s; }
    .tagbtn:hover { box-shadow:0 0 0 2px rgba(0,0,0,.05); }
    .rev { font-size:12px; color:#6E6E6E; display:flex; align-items:center; gap:5px;
           cursor:pointer; user-select:none; margin-left:4px; }
    .note { flex:1; min-width:160px; font-size:12px; padding:6px 10px;
            border:1px solid #EAE2D6; border-radius:8px; color:#432222; }
    .card[data-tag=keep]       { border-left:5px solid #6E6E6E; }
    .card[data-tag=color]      { border-left:5px solid #C9748F; }
    .card[data-tag=name]       { border-left:5px solid #4E7CA1; }
    .card[data-tag=color_name] { border-left:5px solid #8A6BB0; }
    .card[data-tag=cut]        { border-left:5px solid #A24B4B; }
    .card[data-tag=wrong]      { border-left:5px solid #432222; }
    .card[data-tag=missed]     { border-left:5px solid #B8863B; }
    .card[data-tag=other]      { border-left:5px solid #4E8A78; }
    .candhdr { margin-top:14px; font-size:12px; color:#6E6E6E;
               display:flex; align-items:center; gap:10px; flex-wrap:wrap; }
    .swapbtn { font-size:12px; font-weight:600; padding:5px 12px; border:1px solid #4E7CA1;
               background:#fff; color:#4E7CA1; border-radius:8px; cursor:pointer; }
    .swapbtn:hover:not(:disabled) { background:#4E7CA1; color:#fff; }
    .swapbtn:disabled { opacity:.4; cursor:default; }
    .cwrap { cursor:pointer; }
    .cwrap:hover .cimg { box-shadow:0 0 0 2px rgba(78,124,161,.25); }
    .cimg.prop { border:2px solid #4E7CA1; }
    .ovr { color:#4E7CA1; font-weight:600; }
    /* app-only section (icons with no website match) */
    .appsec { margin-top:34px; }
    .appband { display:flex; align-items:center; gap:10px; background:#F1E8F3;
               border-left:5px solid #8A6BB0; padding:11px 15px; margin-bottom:16px; }
    .appttl { font-family:Georgia,'Times New Roman',serif; font-size:16px; color:#5E4680; font-weight:600; }
    .appsub { font-size:12px; color:#8A6BB0; }
    .acard { background:#FBF4F7; border:1px solid #EAD9EC; border-left:5px solid #8A6BB0;
             border-radius:12px; padding:14px 16px; margin-bottom:12px;
             display:flex; align-items:center; gap:16px; }
    .athumb { width:72px; height:72px; border:1px solid #EAD9EC; border-radius:10px;
              background:#fff; object-fit:contain; padding:6px; flex:0 0 auto; }
    .acol { min-width:150px; }
    .apill { display:inline-block; font-size:10px; font-weight:600; color:#6A4E86;
             background:#EEE1F0; padding:2px 8px; border-radius:999px; letter-spacing:.3px; }
    .anm { font-weight:600; font-size:15px; margin-top:6px; }
    .ameta { font-size:12px; color:#6E6E6E; margin-top:2px; }
    .actrl { margin-left:auto; display:flex; align-items:center; gap:8px; flex-wrap:wrap; }
    .abtn { font-size:12px; font-weight:600; padding:6px 15px; border-radius:999px;
            border:1.5px solid; background:transparent; cursor:pointer; }
    .abtn[data-v=new] { color:#C9748F; border-color:#C9748F; }
    .abtn[data-v=new].on { background:#C9748F; color:#fff; }
    .abtn[data-v=other] { color:#4E8A78; border-color:#4E8A78; }
    .abtn[data-v=other].on { background:#4E8A78; color:#fff; }
    .anote { font-size:12px; padding:6px 10px; border:1px solid #EAD9EC; border-radius:8px;
             color:#432222; width:150px; background:#fff; }
    """

    css = css_branded + css_static
    order = ["High confidence", "Needs review", "No match - likely cut"]

    parts = ["<!doctype html><html><head><meta charset='utf-8'>",
             "<title>Abbode Icon Match Review</title><style>", css, "</style></head><body>"]
    parts.append("<h1>Abbode Icon Match Review</h1>")
    parts.append("<div class='sub'>" + esc(meta['shopify_n']) + " old website icons matched against "
                 + esc(meta['app_n']) + " app icons &middot; DINOv2 shape ranking + Claude color confirm"
                 + " &middot; model " + esc(MODEL) + "</div>")
    parts.append("<div class='help'>Tag each pair, then click <b>Download CSV</b>. Buttons start on a "
                 "smart guess (no-match &rarr; Cut; different name &rarr; Name change; otherwise Keep) &mdash; "
                 "override the ones that are wrong, especially color changes. "
                 "<b>Wrong match?</b> Click a better icon in the Candidates row, hit <b>Confirm swap</b>, then pick a "
                 "designation. Use <b>Other</b> for anything out of scope you still need to handle. "
                 "Tags autosave in this browser; the CSV is the real output.</div>")

    # auto-bucket summary (static)
    parts.append("<div class='summary'>")
    for k in order:
        parts.append("<div class='stat'><div class='n'>" + str(counts.get(k, 0)) + "</div>"
                     "<div class='l'>" + esc(k) + "</div></div>")
    parts.append("</div>")

    # sticky toolbar
    parts.append("<div class='toolbar'>")
    parts.append("<div class='counts' id='counts'></div>")
    parts.append("<div class='filters'>")
    for fval, flabel in [("all", "All"), ("unreviewed", "Unreviewed"),
                         ("High confidence", "High conf"), ("Needs review", "Needs review"),
                         ("No match - likely cut", "No match")]:
        on = " class='fon'" if fval == "all" else ""
        parts.append("<button data-f=\"" + esc(fval) + "\"" + on
                     + " onclick=\"setFilter('" + fval.replace("'", "\\'") + "')\">" + esc(flabel) + "</button>")
    parts.append("</div>")
    parts.append("<div class='dl'><span id='progress'></span>"
                 "<button class='dlbtn' onclick='downloadCSV()'>\u2b07 Download CSV</button></div>")
    parts.append("</div>")

    # sections
    for bucket in order:
        group = [r for r in rows if r["bucket"] == bucket]
        if not group:
            continue
        parts.append("<div class='section' data-bucket=\"" + esc(bucket) + "\">")
        parts.append("<h2>" + esc(bucket) + " &middot; " + str(len(group)) + "</h2>")
        for r in group:
            i = r["_id"]
            parts.append("<div class='card' data-id='" + str(i) + "' data-bucket=\"" + esc(bucket) + "\">")
            parts.append("<div class='pair'>")
            parts.append("<img class='thumb' loading='lazy' src='" + r['old_uri'] + "'/>")
            parts.append("<div class='col'><div class='lbl'>Old (website)</div>"
                         "<div class='nm'>" + esc(clean_name(r['old_name'])) + "</div>"
                         "<div class='meta2'>" + esc(r['sku']) + "</div></div>")
            parts.append("<div class='arrow'>&rarr;</div>")
            # New (app) side - always present and JS-updatable (a swap can change or set it)
            new_uri = r['new_uri'] if r['matched'] else ""
            imgstyle = "" if r['matched'] else " style='display:none'"
            parts.append("<img class='thumb newimg' loading='lazy' src='" + new_uri + "'" + imgstyle + "/>")
            parts.append("<div class='col'><div class='lbl'>New (app)"
                         " <span class='ovr' style='display:none'>&middot; manual pick</span></div>"
                         "<div class='nm newname'></div>"
                         "<div class='none nonematch' style='display:none'>&mdash; no match &mdash;</div>"
                         "<div class='meta2 newmeta'></div></div>")
            parts.append("<div class='right'>" + pill(r['conf'], r['bucket'])
                         + "<div class='meta2'>top sim " + ("%.3f" % r['top_sim']) + "</div></div>")
            parts.append("</div>")  # pair
            if r["reason"]:
                parts.append("<div class='reason'>&ldquo;" + esc(r['reason']) + "&rdquo;</div>")
            parts.append("<div class='candhdr'>Candidates &mdash; click a better match, then "
                         "<button class='swapbtn' disabled onclick='confirmSwap(" + str(i) + ")'>Confirm swap</button></div>")
            parts.append("<div class='cands'>")
            for k, c in enumerate(r["cands_view"]):
                parts.append("<div class='cwrap' data-idx='" + str(k) + "' "
                             "onclick='selectCand(" + str(i) + "," + str(k) + ")'>"
                             "<img class='cimg' loading='lazy' src='" + c['uri'] + "'/>"
                             "<div class='csim'>" + ("%.3f" % c['sim']) + "</div>"
                             "<div class='cname'>" + esc(clean_name(c['name'])) + "</div></div>")
            parts.append("</div>")  # cands
            # tagging controls
            parts.append("<div class='controls'>")
            for tval, tlabel, tcolor in TAG_DEFS:
                parts.append("<button class='tagbtn' data-tag='" + tval + "' data-color='" + tcolor + "' "
                             "onclick='setTag(" + str(i) + ",\"" + tval + "\")'>" + esc(tlabel) + "</button>")
            parts.append("<label class='rev'><input type='checkbox' "
                         "onchange='setReviewed(" + str(i) + ",this.checked)'/> Reviewed</label>")
            parts.append("<input class='note' placeholder='note (optional)' "
                         "oninput='setNote(" + str(i) + ",this.value)'/>")
            parts.append("</div>")  # controls
            parts.append("</div>")  # card
        parts.append("</div>")  # section

    # app-only section (icons with no website match = likely new additions)
    if app_only:
        parts.append("<div class='appsec' id='appsec'>")
        parts.append("<div class='appband'><span class='appttl'>App icons with no website match</span>"
                     " <span class='appsub' id='appsub'></span></div>")
        for k, a in enumerate(app_only):
            aid = "app" + str(k)
            meta_extra = (" &middot; " + esc(a['status'])) if a.get('status') else ""
            parts.append("<div class='acard' data-aid='" + aid + "' data-fid=\"" + esc(a['fileid']) + "\">")
            parts.append("<img class='athumb' loading='lazy' src='" + a['uri'] + "'/>")
            parts.append("<div class='acol'><span class='apill'>APP ONLY</span>"
                         "<div class='anm'>" + esc(clean_name(a['name'])) + "</div>"
                         "<div class='ameta'>" + esc(a.get('category', '')) + meta_extra + "</div></div>")
            parts.append("<div class='actrl'>"
                         "<button class='abtn' data-v='new' onclick='setAppTag(\"" + aid + "\",\"new\")'>New</button>"
                         "<button class='abtn' data-v='other' onclick='setAppTag(\"" + aid + "\",\"other\")'>Other</button>"
                         "<label class='rev'><input type='checkbox' "
                         "onchange='setAppReviewed(\"" + aid + "\",this.checked)'/> Reviewed</label>"
                         "<input class='anote' placeholder='note (optional)' "
                         "oninput='setAppNote(\"" + aid + "\",this.value)'/></div>")
            parts.append("</div>")  # acard
        parts.append("</div>")  # appsec

    parts.append("<div class='foot'>Preview only. The full run swaps the baked sample for live Shopify "
                 "pagination and adds the metaobject mapping. App-only cards are computed from the auto "
                 "matches; if you swap a website icon onto one, it drops out of this list automatically.</div>")

    app_js = [{"id": "app" + str(k), "name": a["name"], "fileid": a["fileid"]}
              for k, a in enumerate(app_only)]

    rows_json = json.dumps(row_js, ensure_ascii=True).replace("</", "<\\/")
    colors_json = json.dumps(TAG_COLORS, ensure_ascii=True)

    js_body = r"""
const STORE_KEY = "abbode_match_tags_v1_" + ROWS.length;
let STATE = {};
let PROP = {};      // transient proposed candidate per row (never saved)
let FILTER = "all";

ROWS.forEach(function(r){
  STATE[r.id] = {tag: r.defaultTag, note: "", reviewed: false, matchIdx: r.autoIdx, overridden: false};
});
try {
  var saved = JSON.parse(localStorage.getItem(STORE_KEY) || "{}");
  Object.keys(saved).forEach(function(id){ if (STATE[id]) STATE[id] = Object.assign(STATE[id], saved[id]); });
} catch (e) {}

function save(){ try { localStorage.setItem(STORE_KEY, JSON.stringify(STATE)); } catch (e) {} }
function cardEl(id){ return document.querySelector('.card[data-id="' + id + '"]'); }
function clean(n){ return String(n||"").replace(/\s*Embroidered Icon$/, "").trim(); }

function setTag(id, tag){ STATE[id].tag = tag; STATE[id].reviewed = true; renderRow(id); renderToolbar(); applyFilter(); save(); }
function setNote(id, val){ STATE[id].note = val; save(); }
function setReviewed(id, val){ STATE[id].reviewed = val; renderRow(id); renderToolbar(); applyFilter(); save(); }

function selectCand(id, idx){ PROP[id] = idx; renderRow(id); }
function confirmSwap(id){
  var p = PROP[id];
  if (p === undefined || p === null) return;
  if (p === STATE[id].matchIdx) { PROP[id] = null; renderRow(id); return; }
  STATE[id].matchIdx = p;
  STATE[id].overridden = (p !== ROWS[id].autoIdx);
  PROP[id] = null;
  renderRow(id); renderToolbar(); refreshAppDedup(); applyFilter(); save();
}

function renderRow(id){
  var st = STATE[id]; var el = cardEl(id); if (!el) return;
  var row = ROWS[id];
  el.dataset.tag = st.tag;
  el.querySelectorAll('.tagbtn').forEach(function(b){
    var on = b.dataset.tag === st.tag; var c = b.dataset.color;
    b.style.background = on ? c : "transparent";
    b.style.color = on ? "#fff" : c;
    b.style.borderColor = c;
  });
  // new-match column (reflects current, possibly-swapped, match)
  var mi = st.matchIdx;
  var newimg = el.querySelector('.newimg'), newname = el.querySelector('.newname');
  var newmeta = el.querySelector('.newmeta'), none = el.querySelector('.nonematch'), ovr = el.querySelector('.ovr');
  if (mi >= 0 && row.cands[mi]) {
    var cd = row.cands[mi];
    var cthumb = el.querySelector('.cwrap[data-idx="' + mi + '"] .cimg');
    if (newimg) { if (cthumb) newimg.src = cthumb.src; newimg.style.display = ""; }
    if (newname) newname.textContent = clean(cd.name);
    if (newmeta) newmeta.textContent = (cd.category || "") + (cd.status ? " \u00b7 " + cd.status : "");
    if (none) none.style.display = "none";
  } else {
    if (newimg) newimg.style.display = "none";
    if (newname) newname.textContent = "";
    if (newmeta) newmeta.textContent = "";
    if (none) none.style.display = "";
  }
  if (ovr) ovr.style.display = st.overridden ? "" : "none";
  // candidate rings: pick = current match, prop = proposed
  el.querySelectorAll('.cwrap').forEach(function(w){
    var idx = parseInt(w.dataset.idx, 10); var img = w.querySelector('.cimg');
    img.classList.toggle('pick', idx === st.matchIdx);
    img.classList.toggle('prop', PROP[id] != null && idx === PROP[id] && idx !== st.matchIdx);
  });
  var sb = el.querySelector('.swapbtn');
  if (sb) sb.disabled = !(PROP[id] != null && PROP[id] !== st.matchIdx);
  var chk = el.querySelector('.rev input'); if (chk) chk.checked = !!st.reviewed;
  var ni = el.querySelector('.note'); if (ni && document.activeElement !== ni) ni.value = st.note || "";
}

function renderToolbar(){
  var tags = ["keep","color","name","color_name","cut","wrong","missed","other"];
  var labels = {keep:"Keep", color:"Color", name:"Name", color_name:"Color+Name", cut:"Cut", wrong:"Wrong", missed:"Missed", other:"Other"};
  var cnt = {}; tags.forEach(function(t){ cnt[t] = 0; });
  var reviewed = 0, overridden = 0;
  ROWS.forEach(function(r){ var s = STATE[r.id]; cnt[s.tag] = (cnt[s.tag] || 0) + 1; if (s.reviewed) reviewed++; if (s.overridden) overridden++; });
  document.getElementById('counts').innerHTML = tags.map(function(t){
    return '<span class="cbadge" style="background:' + TAGCOLORS[t] + '">' + labels[t] + ': ' + cnt[t] + '</span>';
  }).join('');
  var total = ROWS.length;
  document.getElementById('progress').textContent =
    "Reviewed " + reviewed + " / " + total + (overridden ? "  \u00b7 " + overridden + " swapped" : "") +
    (reviewed < total ? "  (" + (total - reviewed) + " left)" : "  \u2713 done");
}

function setFilter(f){
  FILTER = f; applyFilter();
  document.querySelectorAll('.filters button').forEach(function(b){ b.classList.toggle('fon', b.dataset.f === f); });
}
function applyFilter(){
  document.querySelectorAll('.card').forEach(function(el){
    var st = STATE[el.dataset.id]; var show = true;
    if (FILTER === "all") show = true;
    else if (FILTER === "unreviewed") show = !st.reviewed;
    else show = (el.dataset.bucket === FILTER);
    el.style.display = show ? "" : "none";
  });
  document.querySelectorAll('.section').forEach(function(sec){
    var any = Array.prototype.slice.call(sec.querySelectorAll('.card')).some(function(c){ return c.style.display !== "none"; });
    sec.style.display = any ? "" : "none";
  });
  var appsec = document.getElementById('appsec');
  if (appsec) {
    var anyApp = false;
    document.querySelectorAll('.acard').forEach(function(el){
      var st = APPSTATE[el.dataset.aid]; var show;
      if (el.dataset.used === "1") show = false;              // swapped upstream -> hidden
      else if (FILTER === "all") show = true;
      else if (FILTER === "unreviewed") show = !st.reviewed;
      else show = false;                                       // bucket filters don't apply to app cards
      el.style.display = show ? "" : "none";
      if (show) anyApp = true;
    });
    appsec.style.display = anyApp ? "" : "none";
  }
}

function csvEsc(v){
  v = (v === null || v === undefined) ? "" : String(v);
  if (/[",\n]/.test(v)) return '"' + v.replace(/"/g, '""') + '"';
  return v;
}
function downloadCSV(){
  var flags = {
    keep:[false,false,false,false,false], color:[true,false,false,false,false], name:[false,true,false,false,false],
    color_name:[true,true,false,false,false], cut:[false,false,true,false,false], wrong:[false,false,false,true,false],
    missed:[false,false,false,false,true], other:[false,false,false,false,false]
  };
  var header = ["shopify_gid","handle","sku","old_name","new_name","matched_png_fileid",
                "designation","color_change","name_change","cut","wrong_match","missed",
                "manual_override","confidence","auto_bucket","reviewed","note"];
  var lines = [header.join(",")];
  ROWS.forEach(function(r){
    var s = STATE[r.id]; var f = flags[s.tag] || [false,false,false,false,false];
    var mi = s.matchIdx;
    var mname = (mi >= 0 && r.cands[mi]) ? r.cands[mi].name : "";
    var mfid = (mi >= 0 && r.cands[mi]) ? r.cands[mi].fileid : "";
    var row = [r.gid, r.handle, r.sku, r.old_name, mname, mfid,
               s.tag, f[0], f[1], f[2], f[3], f[4], s.overridden, r.confidence, r.bucket, s.reviewed, s.note];
    lines.push(row.map(csvEsc).join(","));
  });
  var used = usedFids();
  APPROWS.forEach(function(a){
    if (used[a.fileid]) return;                    // de-dupe: became a manual match upstream
    var s = APPSTATE[a.id];
    var row = ["","","","", a.name, a.fileid, s.tag,
               false,false,false,false,false, false, "", "app-only", s.reviewed, s.note];
    lines.push(row.map(csvEsc).join(","));
  });
  var blob = new Blob([lines.join("\n")], {type: "text/csv;charset=utf-8;"});
  var url = URL.createObjectURL(blob);
  var link = document.createElement("a");
  link.href = url; link.download = "icon_decisions_" + new Date().toISOString().slice(0,10) + ".csv";
  document.body.appendChild(link); link.click(); document.body.removeChild(link);
  setTimeout(function(){ URL.revokeObjectURL(url); }, 1000);
}

// ---- app-only cards (icons with no website match) ----
const APP_KEY = "abbode_match_apptags_v1_" + APPROWS.length;
let APPSTATE = {};
APPROWS.forEach(function(a){ APPSTATE[a.id] = {tag: "new", note: "", reviewed: false}; });
try {
  var sa = JSON.parse(localStorage.getItem(APP_KEY) || "{}");
  Object.keys(sa).forEach(function(id){ if (APPSTATE[id]) APPSTATE[id] = Object.assign(APPSTATE[id], sa[id]); });
} catch (e) {}
function saveApp(){ try { localStorage.setItem(APP_KEY, JSON.stringify(APPSTATE)); } catch (e) {} }
function appEl(id){ return document.querySelector('.acard[data-aid="' + id + '"]'); }
function setAppTag(id, v){ APPSTATE[id].tag = v; APPSTATE[id].reviewed = true; renderAppRow(id); renderAppHeader(); saveApp(); }
function setAppNote(id, v){ APPSTATE[id].note = v; saveApp(); }
function setAppReviewed(id, v){ APPSTATE[id].reviewed = v; renderAppRow(id); renderAppHeader(); saveApp(); }
function renderAppRow(id){
  var st = APPSTATE[id]; var el = appEl(id); if (!el) return;
  el.querySelectorAll('.abtn').forEach(function(b){ b.classList.toggle('on', b.dataset.v === st.tag); });
  var chk = el.querySelector('.rev input'); if (chk) chk.checked = !!st.reviewed;
  var ni = el.querySelector('.anote'); if (ni && document.activeElement !== ni) ni.value = st.note || "";
}
function usedFids(){
  var set = {};
  ROWS.forEach(function(r){ var mi = STATE[r.id].matchIdx; if (mi >= 0 && r.cands[mi]) set[r.cands[mi].fileid] = true; });
  return set;
}
function refreshAppDedup(){
  var used = usedFids();
  document.querySelectorAll('.acard').forEach(function(el){ el.dataset.used = used[el.dataset.fid] ? "1" : ""; });
  renderAppHeader();
}
function renderAppHeader(){
  var used = usedFids();
  var nNew = 0, nOther = 0, rev = 0, tot = 0;
  APPROWS.forEach(function(a){
    if (used[a.fileid]) return;
    tot++; var s = APPSTATE[a.id];
    if (s.tag === "new") nNew++; else nOther++;
    if (s.reviewed) rev++;
  });
  var sub = document.getElementById('appsub');
  if (sub) sub.textContent = tot + " icons \u00b7 New " + nNew + " \u00b7 Other " + nOther + " \u00b7 reviewed " + rev + "/" + tot;
}

ROWS.forEach(function(r){ renderRow(r.id); });
APPROWS.forEach(function(a){ renderAppRow(a.id); });
renderToolbar();
refreshAppDedup();
setFilter("all");
"""

    approws_json = json.dumps(app_js, ensure_ascii=True).replace("</", "<\\/")
    parts.append("<script>\nconst ROWS = " + rows_json + ";\nconst APPROWS = " + approws_json
                 + ";\nconst TAGCOLORS = " + colors_json + ";\n" + js_body + "\n</script>")
    parts.append("</body></html>")
    return "".join(parts)


# --------------------------------------------------------------------------- #
# Main
# --------------------------------------------------------------------------- #
def main():
    t0 = time.time()
    sample = load_shopify_catalog()
    log(f"Loaded {len(sample)} Shopify icons (alphabetical).")

    if not os.environ.get("ANTHROPIC_API_KEY"):
        log("WARNING: ANTHROPIC_API_KEY not set - the vision-confirm step will fail and every "
            "row will fall to 'no match'. Set it for the real result.")

    creds = google_credentials()
    app = read_app_catalog(creds)
    if not app:
        sys.exit("ERROR: No app icons read from the MASTER sheet. Check sheet access / columns.")
    session = make_drive_session(creds)

    # ---- download images (cached) ----
    log("Downloading app icon PNGs (first run only; cached afterwards)...")
    app_ok = []
    for i, a in enumerate(app):
        try:
            a["_bytes"] = fetch_drive_image(session, a["fileid"])
            app_ok.append(a)
        except Exception as e:
            log(f"  skip app '{a['name']}' ({a['fileid']}): {type(e).__name__}")
        if (i + 1) % 100 == 0:
            log(f"  ...{i + 1}/{len(app)}")
    app = app_ok
    log(f"App icons downloaded: {len(app)}")

    log("Downloading Shopify icon PNGs...")
    shop_ok = []
    for i, s in enumerate(sample):
        try:
            s["_bytes"] = fetch_shopify_image(s["image_url"])
            shop_ok.append(s)
        except Exception as e:
            log(f"  skip shopify '{s['old_name']}': {type(e).__name__}")
        if (i + 1) % 100 == 0:
            log(f"  ...{i + 1}/{len(sample)}")
    sample = shop_ok

    # ---- embeddings (cached) ----
    emb = Embedder()
    import numpy as np
    log("Embedding app icons (grayscale DINOv2)...")
    app_vecs = []
    for i, a in enumerate(app):
        app_vecs.append(emb.embed(a["_bytes"]))
        if (i + 1) % 100 == 0:
            log(f"  ...{i + 1}/{len(app)}")
    app_vecs = np.vstack(app_vecs).astype("float32")  # [N, D], L2-normalized

    # pre-decode app color images once (reused as candidate/new-column thumbnails)
    app_color = {}
    for a in app:
        try:
            app_color[a["fileid"]] = on_white(open_rgba(a["_bytes"]))
        except Exception:
            app_color[a["fileid"]] = None

    # ---- rank candidates (sequential, CPU) ----
    log("Ranking candidates...")
    tasks = []
    for si, s in enumerate(sample):
        q = emb.embed(s["_bytes"])
        sims = app_vecs @ q  # cosine (both normalized)
        top_idx = np.argsort(-sims)[:TOP_K]
        cands = [{"name": app[int(j)]["name"], "status": app[int(j)].get("status", ""),
                  "category": app[int(j)].get("category", ""), "fileid": app[int(j)]["fileid"],
                  "img_color": app_color.get(app[int(j)]["fileid"]),
                  "sim": float(sims[j]), "j": int(j)} for j in top_idx]
        try:
            old_color = on_white(open_rgba(s["_bytes"]))
        except Exception:
            old_color = None
        tasks.append({"s": s, "old_color": old_color, "cands": cands,
                      "top_idx": top_idx, "sims": sims})
        if (si + 1) % 100 == 0:
            log(f"  ...ranked {si + 1}/{len(sample)}")

    # ---- confirm with Claude, in parallel ----
    from anthropic import Anthropic
    from concurrent.futures import ThreadPoolExecutor, as_completed
    client = Anthropic()
    log(f"Confirming {len(tasks)} icons with Claude ({CONFIRM_WORKERS} in parallel)...")
    decisions = [None] * len(tasks)
    done = {"n": 0}

    def _do(idx):
        t = tasks[idx]
        d = confirm(client, t["old_color"], t["s"]["old_name"], t["cands"])
        done["n"] += 1
        if done["n"] % 25 == 0 or done["n"] == len(tasks):
            log(f"  ...confirmed {done['n']}/{len(tasks)}")
        return idx, d

    with ThreadPoolExecutor(max_workers=CONFIRM_WORKERS) as ex:
        for fut in as_completed([ex.submit(_do, i) for i in range(len(tasks))]):
            idx, d = fut.result()
            decisions[idx] = d

    # ---- build rows + write thumbnails to disk ----
    log("Writing thumbnails and building report...")
    rows, counts = [], {}
    for idx, t in enumerate(tasks):
        s = t["s"]; cands = t["cands"]; sims = t["sims"]; top_idx = t["top_idx"]
        decision = decisions[idx] or {"match_index": -1, "confidence": "low", "reason": ""}
        di = decision["match_index"]; conf = decision["confidence"]
        if di is None or di < 0 or di >= len(cands):
            matched, bucket = None, "No match - likely cut"
        else:
            matched, bucket = cands[di], ("High confidence" if conf == "high" else "Needs review")
        counts[bucket] = counts.get(bucket, 0) + 1

        old_path = save_img("old_" + safe_key(s.get("sku") or s.get("handle") or str(idx)), t["old_color"])
        cviews = []
        for c in cands:
            cpath = save_img("app_" + safe_key(c["fileid"]), c["img_color"])
            cviews.append({"name": c["name"], "sim": c["sim"], "uri": cpath,
                           "fileid": c["fileid"], "category": c.get("category", ""),
                           "status": c.get("status", ""),
                           "picked": (matched is not None and c["j"] == matched["j"])})
        rows.append({
            "old_name": s["old_name"], "sku": s.get("sku", ""), "gid": s.get("gid", ""),
            "handle": s.get("handle", ""),
            "old_uri": old_path,
            "matched": ({"name": matched["name"], "status": matched["status"],
                         "category": matched["category"], "fileid": matched["fileid"]} if matched else None),
            "new_uri": save_img("app_" + safe_key(matched["fileid"]), matched["img_color"]) if matched else "",
            "conf": conf, "bucket": bucket, "reason": decision.get("reason", ""),
            "top_sim": float(sims[top_idx[0]]),
            "chosen_index": di if matched else -1,
            "cands_view": cviews,
        })

    # ---- app icons that never matched a website icon (= likely new additions) ----
    matched_fids = set(r["matched"]["fileid"] for r in rows if r["matched"])
    app_only_src = [a for a in app if a["fileid"] not in matched_fids]
    app_only_src.sort(key=lambda a: a["name"].lower())
    if APP_ONLY_LIMIT > 0:
        app_only_src = app_only_src[:APP_ONLY_LIMIT]
    app_only = []
    for a in app_only_src:
        app_only.append({"name": a["name"], "fileid": a["fileid"],
                         "category": a.get("category", ""), "status": a.get("status", ""),
                         "uri": save_img("app_" + safe_key(a["fileid"]), app_color.get(a["fileid"]))})
    log(f"App-only icons (no website match): {len(app_only)}")

    # ---- write outputs ----
    meta = {"shopify_n": len(sample), "app_n": len(app)}
    OUT_HTML.write_text(render_html(rows, counts, meta, app_only), encoding="utf-8")

    import csv
    with open(OUT_CSV, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["shopify_gid", "handle", "sku", "old_name", "new_name",
                    "matched_png_fileid", "confidence", "bucket", "top_similarity",
                    "chosen_index", "candidate_names"])
        for r in rows:
            w.writerow([
                r["gid"], r["handle"], r["sku"], r["old_name"],
                r["matched"]["name"] if r["matched"] else "",
                r["matched"]["fileid"] if r["matched"] else "",
                r["conf"], r["bucket"], f"{r['top_sim']:.4f}", r["chosen_index"],
                " | ".join(c["name"] for c in r["cands_view"]),
            ])

    dt = time.time() - t0
    log("-" * 60)
    for k in ["High confidence", "Needs review", "No match - likely cut"]:
        log(f"  {k}: {counts.get(k, 0)}")
    log(f"App-only (new additions): {len(app_only)}")
    log(f"Done in {dt/60:.1f} min.")
    log(f"Report: {OUT_HTML}")


if __name__ == "__main__":
    try:
        main()
    except SystemExit:
        raise
    except Exception:
        traceback.print_exc()
        sys.exit(1)
